
"""
Run this script to generate your marathon training plan Excel file.
Requires: pip install openpyxl
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Training Plan"

# ── Colors ──────────────────────────────────────────────────────────────────
C = {
    "header_bg":  "1F3864",  "header_fg": "FFFFFF",
    "base":       "D9EAD3",  "build":     "FCE5CD",
    "peak":       "F4CCCC",  "taper":     "D0E4F7",
    "lift":       "EAD1DC",  "rest":      "F3F3F3",
    "row_alt":    "FAFAFA",  "border":    "CCCCCC",
}

def fill(hex_color): return PatternFill("solid", fgColor=hex_color)
def bold(size=10, color="000000"): return Font(bold=True, size=size, color=color)
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left(): return Alignment(horizontal="left", vertical="center", wrap_text=True)
def thin_border():
    s = Side(style="thin", color=C["border"])
    return Border(left=s, right=s, top=s, bottom=s)

# ── Column widths ────────────────────────────────────────────────────────────
col_widths = [6, 9, 10, 16, 16, 16, 16, 16, 16, 16, 18, 22]
headers = [
    "Week", "Phase", "Total\nMiles", "Monday", "Tuesday", "Wednesday",
    "Thursday", "Friday", "Saturday", "Sunday", "Lifting\nFocus", "Notes"
]
for i, (w, h) in enumerate(zip(col_widths, headers), 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[1].height = 36

# ── Header row ───────────────────────────────────────────────────────────────
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = fill(C["header_bg"])
    cell.font = bold(10, "FFFFFF")
    cell.alignment = center()
    cell.border = thin_border()

# ── Plan data ─────────────────────────────────────────────────────────────────
# Dates: training starts Mon May 12 2025, race Nov 8 2025 = 26 weeks
# Days: Mon=Lift+Easy, Tue=Quality, Wed=Easy, Thu=Quality/Med, Fri=Lift, Sat=Long, Sun=Rest

plan = [
    # wk, phase,   mi,  Mon,            Tue,                 Wed,          Thu,               Fri,            Sat,           Sun,   Lift focus,        Notes
    (1,  "Base",  28,  "Lift + 4 easy","5 easy",            "5 easy",     "5 easy",          "Lift",         "9 long",      "Rest","Full body / light","Reintroduce running legs. All easy pace (9:30–10:00/mi)."),
    (2,  "Base",  30,  "Lift + 4 easy","5 easy",            "5 easy",     "6 easy",          "Lift",         "10 long",     "Rest","Full body / light","Keep HR conversational. Focus on form."),
    (3,  "Base",  33,  "Lift + 5 easy","6 easy",            "5 easy",     "6 easy",          "Lift",         "11 long",     "Rest","Lower body strength","Introduce strides (4×20 sec) at end of Tue run."),
    (4,  "Base",  28,  "Lift + 4 easy","5 easy",            "4 easy",     "5 easy",          "Lift",         "10 long",     "Rest","Lower body strength","Cutback week. Let body absorb load."),
    (5,  "Base",  36,  "Lift + 5 easy","6 easy + strides",  "5 easy",     "7 med-long",      "Lift",         "13 long",     "Rest","Single-leg / lunges","First med-long Thu (8:45–9:15 pace)."),
    (6,  "Base",  38,  "Lift + 5 easy","6 easy + strides",  "5 easy",     "8 med-long",      "Lift",         "14 long",     "Rest","Single-leg / lunges","Longest base long run. Fuel practice on long run."),
    (7,  "Build", 35,  "Lift + 5 easy","6 w/ 3mi tempo",    "5 easy",     "7 med-long",      "Lift",         "12 long",     "Rest","Glutes / hamstrings","First tempo: 3mi at ~8:00/mi (goal pace). Cutback long."),
    (8,  "Build", 40,  "Lift + 5 easy","7 w/ 4mi tempo",    "5 easy",     "8 med-long",      "Lift",         "15 long",     "Rest","Glutes / hamstrings","Build to 15mi long. Tempo at 7:55–8:05/mi."),
    (9,  "Build", 42,  "Lift + 5 easy","8 w/ 5×800m",       "5 easy",     "8 med-long",      "Lift",         "16 long",     "Rest","Core + stability","Intro intervals: 800m repeats at 7:30/mi w/ 90s jog rest."),
    (10, "Build", 36,  "Lift + 4 easy","6 w/ 3mi tempo",    "5 easy",     "6 med-long",      "Lift",         "13 long",     "Rest","Core + stability","Cutback week. Keep intensity but drop volume."),
    (11, "Build", 44,  "Lift + 5 easy","8 w/ 5×1000m",      "5 easy",     "9 med-long",      "Lift",         "17 long",     "Rest","Single-leg / step-ups","1000m repeats at 7:25/mi. First 17-miler."),
    (12, "Build", 46,  "Lift + 5 easy","9 w/ 6mi tempo",    "5 easy",     "9 med-long",      "Lift",         "18 long",     "Rest","Single-leg / step-ups","Longest tempo block. Practice goal-pace feel."),
    (13, "Build", 40,  "Lift + 5 easy","7 w/ 4mi tempo",    "5 easy",     "7 med-long",      "Lift",         "14 long",     "Rest","Light / deload lift","Cutback. Assess fitness. Half marathon tune-up race optional."),
    (14, "Peak",  48,  "Lift + 6 easy","9 w/ 4×1mi",        "5 easy",     "10 med-long",     "Lift",         "18 long",     "Rest","Maintenance only","Mile repeats at 7:15–7:20. Start practicing race nutrition every long run."),
    (15, "Peak",  50,  "Lift + 6 easy","10 w/ 5×1mi",       "5 easy",     "10 med-long",     "Lift",         "19 long",     "Rest","Maintenance only","Peak mileage week begins. Sleep and nutrition are training."),
    (16, "Peak",  52,  "Lift + 6 easy","10 w/ 7mi tempo",   "5 easy",     "10 med-long",     "Lift",         "20 long",     "Rest","Maintenance only","First 20-miler. Run miles 15–18 at 8:00/mi goal pace."),
    (17, "Peak",  44,  "Lift + 5 easy","8 w/ 4mi tempo",    "5 easy",     "8 med-long",      "Lift",         "16 long",     "Rest","Maintenance only","Cutback. Legs should feel fresher — don't panic."),
    (18, "Peak",  54,  "Lift + 6 easy","10 w/ 6×1mi",       "5 easy",     "10 med-long",     "Lift",         "20 long",     "Rest","Maintenance only","Second 20-miler. Aim for negative split in second half."),
    (19, "Peak",  55,  "Lift + 6 easy","10 w/ 8mi tempo",   "5 easy",     "11 med-long",     "Lift",         "20 long",     "Rest","Maintenance only","Peak fitness week. Longest tempo of the cycle."),
    (20, "Peak",  46,  "Lift + 5 easy","8 w/ 4mi tempo",    "5 easy",     "9 med-long",      "Lift",         "18 long",     "Rest","Maintenance only","Cutback. 18mi with miles 12–16 at goal pace."),
    (21, "Peak",  50,  "Lift + 6 easy","9 w/ 5×1mi",        "5 easy",     "10 med-long",     "Lift",         "19 long",     "Rest","Drop to 1 day lift","Final big training block. Trust the process."),
    (22, "Taper", 40,  "Lift + 5 easy","8 w/ 3mi tempo",    "5 easy",     "8 med-long",      "Lift",         "16 long",     "Rest","1 day lift only","Begin taper. Reduce long run. Keep quality sharp."),
    (23, "Taper", 32,  "Lift + 4 easy","7 w/ 2mi tempo",    "4 easy",     "6 med-long",      "Lift",         "12 long",     "Rest","1 day lift only","Legs may feel heavy or sluggish — totally normal."),
    (24, "Taper", 24,  "Easy 4",       "5 w/ 4×800m",       "4 easy",     "5 med-long",      "Lift",         "8 long",      "Rest","1 day lift only","Drop lifting weight significantly. Feel fresh."),
    (25, "Taper", 16,  "Easy 3",       "4 w/ 3×1mi GP",     "3 easy",     "3 easy",          "Rest",         "4 easy",      "Rest","No lifting","Goal pace miles feel easy? Good. You're ready."),
    (26, "Taper", 6,   "Easy 2",       "3 easy w/ strides", "Rest",       "2 easy + strides","Rest",         "RACE DAY 🏁", "Rest","No lifting","RACE WEEK. Sleep, hydrate, trust your training. Sub-3:30!"),
]

phase_colors = {"Base": C["base"], "Build": C["build"], "Peak": C["peak"], "Taper": C["taper"]}

for r, row in enumerate(plan, 2):
    wk, phase, mi, mon, tue, wed, thu, fri, sat, sun, lift, notes = row
    bg = phase_colors[phase]
    sat_bg = C["taper"] if "RACE" in sat else bg
    values = [wk, phase, mi, mon, tue, wed, thu, fri, sat, sun, lift, notes]
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.border = thin_border()
        cell.alignment = center() if c <= 11 else left()
        if c == 1:   cell.font = bold()
        if c == 2:   cell.fill = fill(bg); cell.font = bold()
        if c == 3:   cell.fill = fill(bg)
        if c in (4,6,8):  cell.fill = fill(C["lift"])      # Mon/Wed/Fri
        if c == 5:   cell.fill = fill(bg)                   # Tue quality
        if c == 7:   cell.fill = fill(bg)                   # Thu
        if c == 9:   cell.fill = fill(sat_bg)               # Sat long
        if c == 10:  cell.fill = fill(C["rest"])            # Sun rest
        if c == 11:  cell.fill = fill(C["lift"])
        if c == 12:  cell.fill = fill(C["row_alt"])
    ws.row_dimensions[r].height = 44

# ── Legend sheet ─────────────────────────────────────────────────────────────
ls = wb.create_sheet("Legend & Pacing")
legend_data = [
    ("PHASE GUIDE", "", ""),
    ("Base (Wks 1–6)", "Build aerobic base, easy running, reintroduce mileage", C["base"]),
    ("Build (Wks 7–13)", "Introduce tempo & intervals, increase long run to 18mi", C["build"]),
    ("Peak (Wks 14–21)", "Highest mileage, race-specific workouts, two 20-milers", C["peak"]),
    ("Taper (Wks 22–26)", "Reduce volume, sharpen speed, arrive fresh on race day", C["taper"]),
    ("", "", ""),
    ("PACE GUIDE (Sub-3:30 = ~8:00/mi avg)", "", ""),
    ("Easy / Recovery", "9:15 – 9:45 / mile", ""),
    ("Long Run", "8:45 – 9:15 / mile (last few miles at 8:00–8:15)", ""),
    ("Med-Long", "8:30 – 9:00 / mile", ""),
    ("Tempo", "7:50 – 8:05 / mile (comfortably hard)", ""),
    ("Intervals (800m/1000m/1mi)", "7:15 – 7:30 / mile", ""),
    ("Goal Race Pace", "8:00 / mile (3:29:48 finish)", ""),
    ("", "", ""),
    ("LIFTING NOTES", "", ""),
    ("Months 1–3", "2x/week: compound lifts at moderate weight (3×8). Squat, DL, RDL, lunges, step-ups.", ""),
    ("Months 4–5", "2x/week: shift to single-leg work, core, hip strength. Reduce squat/DL weight by 20–30%.", ""),
    ("Taper (last 4 wks)", "1x/week only, then drop entirely race week. Protect legs.", ""),
    ("", "", ""),
    ("KEY REMINDERS", "", ""),
    ("Nutrition", "Practice race-day fueling (gels/chews every 45 min) on ALL long runs 14mi+.", ""),
    ("Shoes", "Rotate 2 pairs. Break in race shoes by Week 20.", ""),
    ("Sleep", "8+ hrs is training. Non-negotiable during peak weeks.", ""),
    ("Heart Rate", "80% of runs should feel easy. Resist the urge to run fast on easy days.", ""),
]

ls.column_dimensions["A"].width = 28
ls.column_dimensions["B"].width = 55
for r, (a, b, color) in enumerate(legend_data, 1):
    ca = ls.cell(row=r, column=1, value=a)
    cb = ls.cell(row=r, column=2, value=b)
    ca.alignment = left(); cb.alignment = left()
    if r == 1 or a in ("PACE GUIDE (Sub-3:30 = ~8:00/mi avg)", "LIFTING NOTES", "KEY REMINDERS"):
        ca.font = bold(11, C["header_bg"])
    if color:
        ca.fill = fill(color); cb.fill = fill(color)
    ls.row_dimensions[r].height = 18

fname = "OBX_Marathon_Training_Plan.xlsx"
wb.save(fname)
print(f"✅ Saved: {fname}")
print("Open it in Excel or Google Sheets.")